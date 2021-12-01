from os import read
from functionsForReg import *
from typing import Tuple
import csv
from openpyxl import load_workbook
import pandas as pd
from colorama import init, Fore, Back, Style
from menuFunctions import *

workbook=load_workbook(filename="Users.xlsx")
EmployeesSheet=workbook["Employees"]
ClientsSheet=workbook["Clients"]
VVX=load_workbook(filename="Tariffs.xlsx")
Tariffs=VVX["Tariffs"]

def exit():
    if input("Input 0 if you want to go back to main menu - ")!="0":
        exit()
    return

def user_menu(details):
    while True:
        print('- Menu -\n'
          'Please select the menu number to work with the program\n'
          'My Tariff - 1\n'
          'My balance - 2\n'
          'Subscribe to Tariff - 3\n'
          'Exit the program - 0\n')
        option = int(input('Select an option: '))
        if option == 1:
            showMyTariff(details)
            exit()
            continue
        elif option == 2:
            print(f"Your balance is {details[4]}")
            exit()
            continue
        elif option == 3:
            showTariffs()
            subscribeToNewTariff(details)
            exit()
            continue
        elif option == 0:
            print('Thanks for using our program. Goodbye!')
            break
        else:
            print('Invalid option')
            exit()
            continue
    return

def worker_menu():
    while True:
        print('- Employee menu -\n'
          'Please select the menu number to work with the program\n'
          'List of clients - 1\n'
          'Search - 2\n'
          'Customer history - 3\n'
          'Tariffs - 4\n'
          'Issuing tariffs - 5\n'
          'Sort clients - 6\n'
          'Statistics of clients - 7\n'
          "Add money to Client's balance - 8\n"
          'Exit the program - 0\n')
        option = int(input('Select an option: '))
        if option==1:
            users_list("C")
            exit()
            continue
        elif option==2:
            searchClient()
            exit()
            continue
        elif option==3:
            detailsUser(ClientsSheet)
            exit()
            continue
        elif option==4:
            showTariffs()
            exit()
            continue
        elif option==5:
            viewListOfApplications()
            exit()
            continue
        elif option==6:
            sortClients()
            exit()
            continue
        elif option==7:
            stats()
            exit()
            continue
        elif option==8:
            addBalance()
            exit()
            continue
        elif option == 0:
            print('Thanks for using our program. Goodbye!')
            break

def directorMenu():
    while True:
        print('- Director menu -\n'
          'Worker list - 1\n'
          'Add worker - 2\n'
          'Delete - 3\n'
          'Exit the program - 0')
        option = int(input('Select an option: '))
        if option==1:
            users_list("E")
            exit()
            continue
        elif option==2:
            create_User("E")
            exit()
            continue
        elif option==3:
            detailsUser(EmployeesSheet)
            exit()
            continue
        elif option == 0:
            print('Thanks for using our program. Goodbye!')
            break
