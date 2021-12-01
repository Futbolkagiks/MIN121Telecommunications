from typing import *
from openpyxl import *
import pandas as pd

workbook= load_workbook(filename="Users.xlsx")
EmployeesSheet=workbook["Employees"]
ClientsSheet=workbook["Clients"]

def login_check_2(new_account,type):
    while True:
        LoginCheck=False
        for col in type.iter_rows(min_row=2,values_only=True):
            if new_account[1]==col[2]:
                LoginCheck=True
        if LoginCheck==True:
            print("The login you typed in is already being used, try again")
            new_account[1]=input("Enter another login: ")
        else:
            break

def create_User(type):
    print("Welcome to registration screen!")
    workbook=load_workbook(filename="Users.xlsx")
    EmployeesSheet=workbook["Employees"]
    ClientsSheet=workbook["Clients"]
    if type=="C":
        account_type=ClientsSheet
    elif type=="E":
        account_type=EmployeesSheet
    new_account=[int(len(account_type["A"])),(input("Enter name: ")),(input("Enter login: ")),(input("Enter password: "))]
    login_check_2(new_account,account_type)
    account_type.append(new_account)
    workbook.save("Users.xlsx")
    print("User has been registered")
    return
