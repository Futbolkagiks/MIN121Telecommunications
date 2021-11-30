from functionsForReg import *
from menu import *
from typing import *
from openpyxl import *
from colorama import init, Fore, Back, Style
import pandas as pd
from menuFunctions import *

workbook=load_workbook(filename="Users.xlsx")
EmployeesSheet=workbook["Employees"]
ClientsSheet=workbook["Clients"]
VVX=load_workbook(filename="Tariffs.xlsx")
Tariffs=VVX["Tariffs"]

def account_auth():
    global account_type
    print("Authentication")
    while True:
        chat = input("Type in E if you are an Employee or C for Client: ").upper()
        if chat=="E":
            account_type=EmployeesSheet
            break
        elif chat=="C":
            account_type=ClientsSheet
            break
    CHECK1=False
    CHECK2=False
    while True: #Разобраться с CHECK2
        account=((input("Enter login: ")),(input("Enter password: ")))
        for col in account_type.iter_rows(min_row=2,values_only=True):
            if col[2]==account[0]:
                CHECK1=True
            if col[3]==account[1]:
                global details
                details=col
                CHECK2=True
        if (CHECK1==True) and (CHECK2==True):
            break
        elif CHECK1==False or CHECK2==False:
            print("Try again")
            continue
    print("Success")
    return

print("WELCOME")
while True:
    q = input('To enter the program, write AUTH. If you are a new user, write NEW. If you wish to stop the programm, write EXIT: ').upper()
    if q=="AUTH":
        account_auth()
        break
    elif q=="NEW":
        create_User(ClientsSheet)
        account_auth()
        break
print(Fore.LIGHTYELLOW_EX + "WELCOME")
if account_type==ClientsSheet:
    user_menu(details)
elif account_type==EmployeesSheet:
    worker_menu()