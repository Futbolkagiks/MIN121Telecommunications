from os import read

from pandas.core.algorithms import mode
from functionsForReg import *
from typing import Tuple
import csv
from openpyxl import *
import pandas as pd
from colorama import init, Fore, Back, Style

workbook=load_workbook(filename="Users.xlsx")
EmployeesSheet=workbook["Employees"]
ClientsSheet=workbook["Clients"]
VVX=load_workbook(filename="Tariffs.xlsx")
Tariffs=VVX["Tariffs"]

def showTariffs():
    for row in Tariffs.iter_rows():
        for cell in row:
            print(Fore.CYAN, Style.BRIGHT + f' {cell.value:7}', end=' ')
        print()
    return

def showMyTariff(details):
    while True:
        for col in Tariffs.iter_rows(min_col=1,max_col=3,min_row=2,values_only=True):
            if col[0]==details[5]:
                ActiveT=col[1]
        for col in Tariffs.iter_rows(min_col=1,max_col=3,min_row=2,values_only=True):
            if col[0]==details[6]:
                PreviousT=col[1]
        print(f"Your active Tariff is {ActiveT}")
        print(f"Your previous Tariff was {PreviousT}")
        return

def users_list(XD):
    if XD=="C":
        type=ClientsSheet
    elif XD=="E":
        type=EmployeesSheet
    for col in type.iter_rows(values_only=True):
        info=[col[0],col[1],col[2],col[4]]
        print(Fore.LIGHTGREEN_EX, Style.BRIGHT + f' {info}', end=' \n')
    return

def searchClient():
    found=False
    name=str(input("Please input who you want to search: "))
    while True:
        for col in ClientsSheet.iter_rows(min_row=2,values_only=True):
            if (name.upper() in col[1].upper())==True:
                found=True
                print("ID     NAME     LOGIN      BALANCE")
                print(col[0],col[1],col[2],col[4])
        if found==False:
            print("Your input was wrong. Try again")
            continue
        elif found==True:
            break
    return

def detailsUser(type):
    idUser=input("Please enter ID of the USER: ")
    for col in type.iter_rows(min_row=2,values_only=True):
        if int(idUser)==int(col[0]):
            print(col[:3],col[4:])
    return

def sortClients():
    print(
    'ID - 1\n'
    'Name - 2\n'
    'Age - 3\n'
    'City - 4\n')
    c = input("Enter by what you wish to sort clients: ")
    b = pd.read_excel('Users.xlsx',sheet_name="Clients")
    lst = []
    for i in range(len(b["Id"])):
        lst.append([b["Id"][i], b["Name"][i], b["Age"][i], b["City"][i]])
    if c == '1':
        lst = sorted(lst, key=lambda x: x[0])
        for i in lst:
            print(i)
    elif c == '2':
        lst = sorted(lst, key=lambda x: x[1])
        for i in lst:
            print(i)
    elif c == '3':
        lst = sorted(lst, key=lambda x: x[2])
        for i in lst:
            print(i)
    elif c== "4":
        lst = sorted(lst, key=lambda x: x[3])
        for i in lst:
            print(i)
    return

def stats():
    age = []
    local = []
    b = pd.read_excel('Users.xlsx',sheet_name="Clients")
    for i in range(len(b['Age'])):
        age.append(b['Age'][i])
        local.append(b['City'][i])
    age1 = set(age)
    local1 = set(local)
    print('Stats by City')
    for i in local1:
        print(local.count(i), i)
    print('Stats by Age')
    for i in age1:
        print(age.count(i), i)

def addBalance():
    idUser=input("Enter the ID of a client: ")
    count=1
    for col in ClientsSheet.iter_rows(min_row=2,values_only=True):
        count+=1
        if int(idUser)==int(col[0]):
            ClientsSheet[f"E{count}"]=int(col[4])+int(input("How much?: "))
            workbook.save("Users.xlsx")
    print("Money has been added to the balance of a client")

def subscribeToNewTariff(details):
    idTariff=input("Enter the ID of Tariff you want to use: ")
    thefile=pd.DataFrame([details[0],idTariff])
    thefile.to_csv("Applications.csv",mode='a',header=False)
    print("Your application has been submitted")
    return

def ApplicationSubFunction1(id):
    for col in ClientsSheet.iter_rows(min_row=2,values_only=True):
        if int(id)==int(col[0]):
            return col

def ApplicationSubFunction2(id):
    for row in Tariffs.iter_rows(min_row=2,values_only=True):
        if row[0]==id:
            return row

def viewListOfApplications():
    thefile=pd.read_csv("Applications.csv")
    for row in range(len(thefile['clientId'])):
        q=[(thefile['clientId'][row]),(thefile['tariffId'][row])]
        infoClient=ApplicationSubFunction1(q[0])
        infoTariff=ApplicationSubFunction2(q[1])
        print(f"User {infoClient[1]} wishes to use Tariff {infoTariff[1]}")
    
def addInfoToClient(details):
    print("Extra INFO screen")
    add_info=[input("Enter the city where you live: "), input("Enter your age: ")]
    count=1
    for col in ClientsSheet.iter_rows(min_row=2,values_only=True):
        count+=1
        if int(details[0])==int(col[0]):
            ClientsSheet[f"H{count}"]=add_info[1]
            ClientsSheet[f"I{count}"]=add_info[0]
            ClientsSheet[f"E{count}"]=0
            workbook.save("Users.xlsx")
    print("Extra informations has been saved")